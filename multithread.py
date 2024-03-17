import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from concurrent.futures import ThreadPoolExecutor

def get_links_from_xlsx():
    df = pd.read_excel('part_19.xlsx')
    return df

def array_to_xlsx(arr):
    if arr is None:
        print("Error: Array is None.")
        return
    
    output_file = "extracted_data.xlsx"
    if os.path.isfile(output_file):
        wb = load_workbook(output_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
    
    ws.append(arr)
    
    wb.save(output_file)

def save_to_csv(df):
    file_path = "extracted_data.csv"
    
    if os.path.isfile(file_path):
        past_df = pd.read_csv(file_path)
        combined_df = pd.concat([past_df, df], ignore_index=True)
        combined_df.to_csv(file_path, index=False)
        print(f"Data added to '{file_path}'")
    else:
        df.to_csv(file_path, index=False)
        print(f"Data stored in a new file '{file_path}'")

def scrape_website(url):
    response = requests.get(url, verify=False)

    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        
        cards = soup.find_all('h3', class_='cards-title')
        extracted_data = []
        for card in cards:
            detail_link = card.find('a')['href']
            detail_url = f"https://canada.constructconnect.com{detail_link}"
            extracted_data.append(scrape_detail_page(detail_url))

        return pd.concat(extracted_data, ignore_index=True)
    else:
        print(f"Failed to fetch webpage: {url}")

def scrape_detail_page(detail_url):
    response = requests.get(detail_url)

    extracted_data = []
    
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        
        sec_contents = soup.find_all('div', class_='content-left')
        data = {}
        for sec_content in sec_contents:
            for index, content in enumerate(sec_content.find_all('section', class_='content')):
                if index == 0:
                    title_tag = content.find('h3')
                    if title_tag:
                        title = title_tag.text.strip()
                        section_text = content.get_text().strip()
                        form_id = section_text.split('\n')[-1].strip()
                        data[title] = form_id
                elif index == 1:
                    pb_title_tag = content.find('h3')
                    pb_time_tag = content.find('time')
                    if pb_title_tag and pb_time_tag:
                        data[pb_title_tag.text.strip()] = pb_time_tag.text.strip()
                elif index == 3:
                    dt_tag = content.find('dt')
                    if dt_tag is not None:
                        data["address_A"] = dt_tag.text.strip()
                elif index == 4:
                    dt_tag = content.find('dt')
                    if dt_tag is not None:
                        data["address_B"] = dt_tag.text.strip()
                else:
                    dt_tags = content.find_all('dt')
                    dd_tags = content.find_all('dd')
                    for dt_tag, dd_tag in zip(dt_tags, dd_tags):
                        data[dt_tag.text.strip()] = dd_tag.text.strip()

        file_links = soup.find_all('p', class_='')
        for index, file_link in enumerate(file_links):
            link = file_link.find('a')
            if link:
                data[f"Link{index}"]= link['href']

        data["src_url"] = detail_url

        if data:
            extracted_data.append(data)

    return pd.DataFrame(extracted_data)

def main():
    df_links = get_links_from_xlsx()
    arr_links = df_links.iloc[:, 0].values

    with ThreadPoolExecutor(max_workers=7) as executor:
        for result in executor.map(scrape_website, arr_links):
            save_to_csv(result)
main()