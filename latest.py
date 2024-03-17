import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from concurrent.futures import ThreadPoolExecutor
import requests.exceptions

requests.urllib3.disable_warnings(requests.urllib3.exceptions.InsecureRequestWarning)

def get_links_from_xlsx():  
    df = pd.read_excel('part_38.xlsx')
    return df

def array_to_xlsx(arr):
    if arr is None:
        print("Error: Array is None.")
        return
    
    output_file = "extracted_data1.xlsx"
    if os.path.isfile(output_file):
        wb = load_workbook(output_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
    
    ws.append(arr)
    
    wb.save(output_file)

def save_to_csv(df):
    file_path = "extracted_data38.csv"
    
    if os.path.isfile(file_path):
        past_df = pd.read_csv(file_path)
        combined_df = pd.concat([past_df, df], ignore_index=True)
        combined_df.to_csv(file_path, index=False)
        print(f"Data added to '{file_path}'")
    else:
        df.to_csv(file_path, index=False)
        print(f"Data stored in a new file '{file_path}'")

def scrape_website(url):
    response = requests.get(url, verify=False)

    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        
        cards = soup.find_all('h3', class_='cards-title')
        df = pd.DataFrame()
        
        for card in cards:
            detail_link = card.find('a')['href']
            
            extracted_data = scrape_detail_page(f"https://canada.constructconnect.com{detail_link}")
            df = df._append(extracted_data, ignore_index=True)

        return df
    else:
        print("Failed to fetch webpage")


def scrape_detail_page(detail_url):
    response = requests.get(detail_url, verify=False)

    extracted_data = []
    
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        
        sec_contents = soup.find_all('div', class_='content-left')
        data = {}
        form_id_flg = True
        for sec_content in sec_contents:
            for index, content in enumerate(sec_content.find_all('section', class_='content')):
                if index == 0:
                    title_tag = content.find('h3')
                    if title_tag:
                        title = title_tag.text.strip()
                        section_text = content.get_text().strip()
                        form_id = section_text.split('\n')[-1].strip()
                        data[title] = form_id
                        if form_id == "Form ID":
                            form_id_flg = False
                elif form_id_flg:
                    if index == 1 and form_id_flg:
                        pb_title_tag = content.find('h3')
                        pb_time_tag = content.find('time')
                        if pb_title_tag and pb_time_tag:
                            data[pb_title_tag.text.strip()] = pb_time_tag.text.strip()
                    elif index == 3:
                        dt_tag = content.find('dt')
                        if dt_tag is not None:
                            data["Location of premises"] = dt_tag.text.strip()
                    elif index == 4:
                        dt_tag = content.find('dt')
                        if dt_tag is not None:
                            data["Location of premises"] += f", {dt_tag.text.strip()}"
                    else:
                        dt_tags = content.find_all('dt')
                        dd_tags = content.find_all('dd')
                        for dt_tag, dd_tag in zip(dt_tags, dd_tags):
                            if dt_tag.text.strip() not in data:
                                data[dt_tag.text.strip()] = dd_tag.text.strip()
                            else:
                                data[dt_tag.text.strip()] += f", {dd_tag.text.strip()}"
                else:
                    if index == 1 and form_id_flg:
                        pb_title_tag = content.find('h3')
                        pb_time_tag = content.find('time')
                        if pb_title_tag and pb_time_tag:
                            data[pb_title_tag.text.strip()] = pb_time_tag.text.strip()
                    elif index == 2:
                        lp_title = content.find('h3')
                        lp_title_txt = lp_title.text.strip()
                        ct_txt1 = content.find('h4')
                        ct_txt1_txt = ct_txt1.text.strip()
                        ct_txt2 = content.find('p', class_="print-hidden")
                        ct_txt2_txt = ct_txt2.text.strip()
                        data[lp_title_txt] = f"{ct_txt1_txt}, {ct_txt2_txt}"
                    elif index == 3:
                        ct_title = content.find('h3')
                        ct_title_txt = ct_title.text.strip()
                        ct_ctt_txt1 = content.find('p')
                        ct_ctt_txt1_txt = ct_ctt_txt1.text.strip()
                        dt_tags = content.find_all('dt')
                        dd_tags = content.find_all('dd')
                        t_txt = f"{ct_ctt_txt1_txt}"
                        for dt_tag, dd_tag in zip(dt_tags, dd_tags):
                            if dt_tag.text.strip() not in data:
                                data[dt_tag.text.strip()] = dd_tag.text.strip()
                            else:
                                data[dt_tag.text.strip()] += f", {dd_tag.text.strip()}"
                        data[ct_title_txt] = t_txt
                    else:
                        ct_title = content.find('h3')
                        ct_title_txt = ct_title.text.strip()
                        dt_tags = content.find_all('dt')
                        dd_tags = content.find_all('dd')
                        for dt_tag, dd_tag in zip(dt_tags, dd_tags):
                            if dt_tag.text.strip() not in data:
                                data[dt_tag.text.strip()] = dd_tag.text.strip()
                            else:
                                data[dt_tag.text.strip()] += f", {dd_tag.text.strip()}"
                        data[ct_title_txt] = ""

        file_links = soup.find_all('p', class_='')
        for index, file_link in enumerate(file_links):
            link = file_link.find('a')
            if link:
                data[f"Link{index}"]= link['href']

        data["src_url"] = detail_url

        extracted_data.append(data)

        df = pd.DataFrame(extracted_data)
        return df
    else:
        print("Failed to fetch webpage")

def main():
    df_links = get_links_from_xlsx()
    arr_links = df_links.iloc[:, 0].values

    with ThreadPoolExecutor(max_workers=7) as executor:
        for result in executor.map(scrape_website, arr_links):
            save_to_csv(result)


main()